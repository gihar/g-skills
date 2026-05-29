"""Merge updated Jira data into an existing HR-backlog xlsx.

Layout (2026+):
  Row 1: empty
  Row 2: «Итого задач» (F) + SUBTOTAL count (G); «Итого часов» (H) + SUBTOTAL sum (I)
  Row 3: headers
  Row 4: empty separator
  Row 5+: data

Columns:
  A=1  №CAB           (key)
  B=2  ФН             (MANUAL — not touched)
  C=3  Страна         (MANUAL — not touched)
  D=4  Статус         (priority)
  E=5  Название        (summary)
  F=6  Статус          (workflow status)         <-- SKIP-trigger if «Выполнено»/«Отмена»/«План»
  G=7  Инициатор       (customer)
  H=8  Система         (MANUAL — not touched)
  I=9  План в ч/д
  J=10 Комментарий
  K=11 Фактическая дата выхода в тест    (max current customfield_11944 across active FI)
  L=12 Плановая дата выхода в тест       (first changelog set of customfield_11944 while FI in Запланирована)
  M=13 Фактическая дата завершения       (last changelog transition to Выполнено)
  N=14 Плановая дата завершения          (max current customfield_10311 across active FI)
  O=15 Дата обновления                   (today() OR label)

Inputs:
  --input PATH        existing xlsx (read-only source)
  --cab-json PATH     distilled CAB issues (dict CAB-XXXX -> {...})
  --fi-json PATH      distilled FI issues (dict FI-XXXX -> {...})
  --non-jql PATH      optional: dict CAB-XXXX -> {issuetype, status_category, status_name, resolutiondate}
  --output PATH       output xlsx
  --cutoff-date       ISO date (default 2026-01-01)
  --today             ISO date (default today)

Behavior:
  * Skip-statuses {«Выполнено», «Отмена», «План»}: row is not touched at all
    (data, dates, label — nothing). Counted as `skipped`.
  * Manual columns B, C, H — never touched.
  * For all other rows present in CAB JQL: overwrite data + stamp O = today.
  * For rows missing from CAB JQL: stamp label in O (placeholder/invalid/non-CR/done-before/…).
  * New CAB keys appended at the end.
  * SUBTOTAL formulas rewritten (G + I) over current data range.
"""

from __future__ import annotations
import argparse
import json
import re
from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path
from typing import Any

import openpyxl

CAB_KEY_RE = re.compile(r"^CAB-(\d+)$")

SKIP_STATUSES = {"Выполнено", "Отмена", "Отменено", "План"}

COL = {
    "key":       1,
    "fn":        2,
    "country":   3,
    "priority":  4,
    "summary":   5,
    "status":    6,
    "customer":  7,
    "system":    8,
    "plan":      9,
    "comment":  10,
    "fact_test": 11,
    "plan_test": 12,
    "fact_done": 13,
    "plan_due":  14,
    "update":   15,
}

HEADERS = {
    "key":       "№CAB",
    "fn":        "ФН",
    "country":   "Страна",
    "priority":  "Статус",
    "summary":   "Название",
    "status":    "Статус",
    "customer":  "Инициатор",
    "system":    "Система",
    "plan":      "План в ч/д",
    "comment":   "Комментарий",
    "fact_test": "Фактическая дата выхода в тест",
    "plan_test": "Плановая дата выхода в тест",
    "fact_done": "Фактическая дата завершения",
    "plan_due":  "Плановая дата завершения",
    "update":    "Дата обновления",
}

HEADER_ROW = 3
SEPARATOR_ROW = 4
FIRST_DATA_ROW = 5


def is_valid_cab_key(s: str) -> bool:
    return bool(s) and bool(CAB_KEY_RE.fullmatch(s.strip()))


def has_cyrillic_in_key_prefix(s: str) -> bool:
    if not s:
        return False
    head = s.split("-", 1)[0]
    return any("А" <= ch <= "я" or ch in "Ёё" for ch in head)


def parse_iso_date(s: Any) -> date | None:
    if not s:
        return None
    if isinstance(s, datetime):
        return s.date()
    if isinstance(s, date):
        return s
    txt = str(s)
    try:
        return datetime.strptime(txt[:10], "%Y-%m-%d").date()
    except ValueError:
        return None


def fmt_date(d: date | None) -> str:
    return d.strftime("%Y-%m-%d") if d else ""


def plan_value(plan_zup: Any, plan_e: Any) -> Any:
    if plan_zup is None and plan_e is None:
        return ""
    a = plan_zup if isinstance(plan_zup, (int, float)) else 0
    b = plan_e if isinstance(plan_e, (int, float)) else 0
    return a + b


def best_iso_date(dates: list[Any]) -> str:
    """Return max ISO date string among inputs (parses+formats). Empty when no valid dates."""
    parsed = [parse_iso_date(d) for d in dates if d]
    parsed = [p for p in parsed if p]
    if not parsed:
        return ""
    return max(parsed).isoformat()


def cutoff_label_for_missing(year: int, info: dict) -> str:
    itype = (info.get("issuetype") or "").strip()
    status_cat = (info.get("status_category") or "").strip().lower()
    status_name = (info.get("status_name") or "").strip()
    if itype and itype != "Change Request":
        return f"Не Change Request (тип {itype})"
    if status_cat == "done":
        if "отмен" in status_name.lower():
            return f"Отменена до {year}"
        return f"Выполнена до {year}"
    return "Не в JQL"


@dataclass
class SheetCtx:
    ws: Any
    header_row: int = HEADER_ROW
    first_data_row: int = FIRST_DATA_ROW


def find_sheet(wb) -> SheetCtx | None:
    for ws in wb.worksheets:
        try:
            v = ws.cell(row=HEADER_ROW, column=COL["key"]).value
            if not v:
                continue
            normalized = (
                str(v).upper()
                .replace("С", "C").replace("А", "A").replace("В", "B")
            )
            if "CAB" in normalized:
                return SheetCtx(ws=ws)
        except Exception:
            continue
    if wb.worksheets:
        return SheetCtx(ws=wb.worksheets[0])
    return None


def get_existing_key_rows(ctx: SheetCtx) -> list[tuple[str, int]]:
    """Return [(key, row), ...] preserving duplicates (e.g. multiple 'нет' rows)."""
    ws = ctx.ws
    out: list[tuple[str, int]] = []
    row = ctx.first_data_row
    empties = 0
    while row < 6000:
        v = ws.cell(row=row, column=COL["key"]).value
        if v is None or str(v).strip() == "":
            empties += 1
            if empties >= 20:
                break
        else:
            empties = 0
            out.append((str(v).strip(), row))
        row += 1
    return out


def ensure_headers(ctx: SheetCtx):
    """Fill missing header cells (overwrite blanks only — keep existing labels)."""
    for key in ("fact_test", "plan_test", "fact_done", "plan_due", "update"):
        cell = ctx.ws.cell(row=ctx.header_row, column=COL[key])
        if not cell.value:
            cell.value = HEADERS[key]


def collect_fi_for_cab(cab_data: dict, fi_data: dict[str, dict]) -> list[dict]:
    return [fi_data[k] for k in (cab_data.get("fi_keys") or []) if fi_data.get(k)]


def write_row_data(
    ctx: SheetCtx, row: int, *,
    cab_data: dict, fi_data: dict[str, dict],
    cutoff: date, today_str: str,
):
    ws = ctx.ws
    ws.cell(row=row, column=COL["priority"]).value = cab_data.get("priority", "")
    ws.cell(row=row, column=COL["summary"]).value = cab_data.get("summary", "")
    ws.cell(row=row, column=COL["status"]).value = cab_data.get("status_name", "")
    ws.cell(row=row, column=COL["customer"]).value = cab_data.get("customer") or ""
    ws.cell(row=row, column=COL["plan"]).value = plan_value(
        cab_data.get("plan_zup"), cab_data.get("plan_e"),
    )
    ws.cell(row=row, column=COL["comment"]).value = cab_data.get("comment", "")

    fis = collect_fi_for_cab(cab_data, fi_data)

    # fact_test_date: max(current customfield_11944) across active FIs
    active = [
        fi for fi in fis
        if fi.get("status_category") != "done"
        or (parse_iso_date(fi.get("resolutiondate")) or date.min) >= cutoff
    ]
    ws.cell(row=row, column=COL["fact_test"]).value = best_iso_date(
        [fi.get("fact_test_date") for fi in active]
    )

    # plan_test_date: max across ALL FIs (historical fact, no active-filter)
    ws.cell(row=row, column=COL["plan_test"]).value = best_iso_date(
        [fi.get("plan_test_date") for fi in fis]
    )

    # fact_done_date: max across ALL FIs (last transition to Выполнено from changelog)
    ws.cell(row=row, column=COL["fact_done"]).value = best_iso_date(
        [fi.get("fact_done_date") for fi in fis]
    )

    # plan_due_date: max(current customfield_10311) across active FIs;
    # if none active have a value, fall back to all FIs.
    pd_active = best_iso_date([fi.get("plan_due_date") for fi in active])
    ws.cell(row=row, column=COL["plan_due"]).value = (
        pd_active or best_iso_date([fi.get("plan_due_date") for fi in fis])
    )

    ws.cell(row=row, column=COL["update"]).value = today_str


def write_label_only(ctx: SheetCtx, row: int, label: str):
    ctx.ws.cell(row=row, column=COL["update"]).value = label


def append_new_row(
    ctx: SheetCtx, key: str, *,
    cab_data: dict, fi_data: dict[str, dict],
    cutoff: date, today_str: str,
) -> int:
    ws = ctx.ws
    last = ctx.first_data_row - 1
    row = ctx.first_data_row
    while row < 6000:
        v = ws.cell(row=row, column=COL["key"]).value
        if v is not None and str(v).strip():
            last = row
        row += 1
        if row - last > 60:
            break
    new_row = last + 1
    ws.cell(row=new_row, column=COL["key"]).value = key
    write_row_data(
        ctx, new_row,
        cab_data=cab_data, fi_data=fi_data,
        cutoff=cutoff, today_str=today_str,
    )
    return new_row


def rewrite_subtotals(ctx: SheetCtx):
    ws = ctx.ws
    last = ctx.first_data_row - 1
    row = ctx.first_data_row
    while row < 6000:
        v = ws.cell(row=row, column=COL["key"]).value
        if v is not None and str(v).strip():
            last = row
        row += 1
        if row - last > 100:
            break
    if last < ctx.first_data_row:
        return
    ws.cell(row=2, column=7).value = f"=SUBTOTAL(3,G{ctx.first_data_row}:G{last})"
    ws.cell(row=2, column=9).value = f"=SUBTOTAL(9,I{ctx.first_data_row}:I{last})"


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--input", required=True)
    ap.add_argument("--cab-json", required=True)
    ap.add_argument("--fi-json", required=True)
    ap.add_argument("--non-jql", default=None)
    ap.add_argument("--output", required=True)
    ap.add_argument("--cutoff-date", default="2026-01-01")
    ap.add_argument("--today", default=None)
    args = ap.parse_args()

    cutoff = datetime.strptime(args.cutoff_date, "%Y-%m-%d").date()
    today = (
        datetime.strptime(args.today, "%Y-%m-%d").date()
        if args.today else date.today()
    )
    today_str = today.isoformat()
    year = cutoff.year

    with open(args.cab_json) as f:
        cab = json.load(f)
    with open(args.fi_json) as f:
        fi = json.load(f)
    non_jql = {}
    if args.non_jql and Path(args.non_jql).exists():
        with open(args.non_jql) as f:
            non_jql = json.load(f)

    wb = openpyxl.load_workbook(args.input)
    ctx = find_sheet(wb)
    if ctx is None:
        raise SystemExit("No usable sheet found")
    ensure_headers(ctx)

    key_rows = get_existing_key_rows(ctx)
    keys_in_file = {k for k, _ in key_rows}

    updated, added, labeled, skipped = 0, 0, 0, 0
    file_keys_not_in_jql: list[str] = []
    skipped_status_breakdown: dict[str, int] = {}
    label_counts: dict[str, int] = {
        "placeholders": 0,
        "invalid_cyrillic": 0,
        "invalid_other": 0,
        "non_cr": 0,
        "cancelled_before": 0,
        "done_before": 0,
        "not_in_jql": 0,
    }

    for key, row in key_rows:
        # SKIP-statuses checked FIRST, before any cab-lookup / label-logic.
        # Per spec: задачи в статусах «Выполнено», «Отмена», «План» НЕ ОБНОВЛЯЮТСЯ
        # никогда — независимо от того, есть ли они в JQL или нет.
        status_in_file = ctx.ws.cell(row=row, column=COL["status"]).value
        status_norm = str(status_in_file).strip() if status_in_file else ""
        if status_norm in SKIP_STATUSES:
            skipped += 1
            skipped_status_breakdown[status_norm] = (
                skipped_status_breakdown.get(status_norm, 0) + 1
            )
            continue  # do not touch this row at all

        if key in cab:
            write_row_data(
                ctx, row,
                cab_data=cab[key], fi_data=fi,
                cutoff=cutoff, today_str=today_str,
            )
            updated += 1
        else:
            k_low = key.lower().strip()
            if k_low in {"нет", "-", "—", "tbd", "?"}:
                write_label_only(ctx, row, "Плейсхолдер (нет CAB)")
                label_counts["placeholders"] += 1
            elif has_cyrillic_in_key_prefix(key):
                write_label_only(ctx, row, "Некорректный ключ (кириллица)")
                label_counts["invalid_cyrillic"] += 1
            elif not is_valid_cab_key(key):
                write_label_only(ctx, row, "Некорректная запись")
                label_counts["invalid_other"] += 1
            elif key in non_jql:
                lbl = cutoff_label_for_missing(year, non_jql[key])
                write_label_only(ctx, row, lbl)
                if lbl.startswith("Не Change Request"):
                    label_counts["non_cr"] += 1
                elif lbl.startswith("Отменена"):
                    label_counts["cancelled_before"] += 1
                elif lbl.startswith("Выполнена"):
                    label_counts["done_before"] += 1
                else:
                    label_counts["not_in_jql"] += 1
                    file_keys_not_in_jql.append(key)
            else:
                write_label_only(ctx, row, "Не в JQL")
                label_counts["not_in_jql"] += 1
                file_keys_not_in_jql.append(key)
            labeled += 1

    for key in cab:
        if key not in keys_in_file:
            append_new_row(
                ctx, key,
                cab_data=cab[key], fi_data=fi,
                cutoff=cutoff, today_str=today_str,
            )
            added += 1

    rewrite_subtotals(ctx)
    Path(args.output).parent.mkdir(parents=True, exist_ok=True)
    wb.save(args.output)

    stats = {
        "today": today_str,
        "cutoff_date": args.cutoff_date,
        "input": args.input,
        "output": args.output,
        "updated": updated,
        "added": added,
        "labeled": labeled,
        "skipped": skipped,
        "skipped_status_breakdown": skipped_status_breakdown,
        "label_breakdown": label_counts,
        "file_keys_not_in_jql": file_keys_not_in_jql,
        "cab_total": len(cab),
        "fi_total": len(fi),
    }
    stats_path = Path(args.output).with_name("merge_stats.json")
    stats_path.write_text(json.dumps(stats, ensure_ascii=False, indent=2))

    print(f"updated={updated} added={added} labeled={labeled} skipped={skipped}")
    print(
        f"  skipped by status: {skipped_status_breakdown}\n"
        f"  placeholders={label_counts['placeholders']} "
        f"done_before={label_counts['done_before']} "
        f"cancelled_before={label_counts['cancelled_before']} "
        f"non_cr={label_counts['non_cr']} "
        f"invalid_cyrillic={label_counts['invalid_cyrillic']} "
        f"invalid_other={label_counts['invalid_other']} "
        f"not_in_jql={label_counts['not_in_jql']}"
    )
    print(f"file_keys_marked_'Не в JQL' (need lookup): {file_keys_not_in_jql}")
    print(f"stats  -> {stats_path}")
    print(f"output -> {args.output}")


if __name__ == "__main__":
    main()
