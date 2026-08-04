#!/usr/bin/env python3
"""Generate a structured Crocotime monthly worked-time report in XLSX."""

from __future__ import annotations

import argparse
import json
import os
import re
import ssl
import subprocess
import sys
import urllib.error
import urllib.request
from collections import defaultdict
from dataclasses import dataclass
from datetime import date, datetime, timedelta
from pathlib import Path
from typing import Any
from zoneinfo import ZoneInfo

from openpyxl import Workbook, load_workbook
from openpyxl.formatting.rule import DataBarRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

MONTHS_RU = (
    "",
    "Январь", "Февраль", "Март", "Апрель", "Май", "Июнь",
    "Июль", "Август", "Сентябрь", "Октябрь", "Ноябрь", "Декабрь",
)
WEEKDAYS_RU = ("Пн", "Вт", "Ср", "Чт", "Пт", "Сб", "Вс")
INVALID_SHEET_CHARS = re.compile(r"[\\/*?:\[\]]")

HEADER_FILL = PatternFill("solid", fgColor="D9EAF7")
HEADER_FONT = Font(name="Arial", size=10, color="2E5E8C", bold=True)
GROUP_FILL = PatternFill("solid", fgColor="EDF3F8")
GREEN_FILL = PatternFill("solid", fgColor="2BA323")
LIGHT_GREEN_FILL = PatternFill("solid", fgColor="A9DFA4")
GRAY_FILL = PatternFill("solid", fgColor="D1D3D8")
RED_FILL = PatternFill("solid", fgColor="E56B6F")
WHITE_FILL = PatternFill("solid", fgColor="FFFFFF")
THIN_BLUE = Side(style="thin", color="D6E2EC")
ROW_BORDER = Border(bottom=THIN_BLUE)
DURATION_FORMAT = '[h]" ч "mm" мин"'
SIGNED_DURATION_FORMAT = '[Green]+[h]" ч "mm" мин";[Red]-[h]" ч "mm" мин";0" мин"'
TIME_FORMAT = "hh:mm"
PERCENT_FORMAT = "0%"


class CrocotimeError(RuntimeError):
    pass


class CrocotimeClient:
    def __init__(self, base_url: str, token: str, app_version: str | None = None, timeout: int = 60):
        self.base_url = base_url.rstrip("/") + "/"
        self.token = token
        self.app_version = app_version
        self.timeout = timeout
        # Некоторые инсталляции Crocotime принимают только легаси-шифры TLS 1.2
        # (SECLEVEL=1); с настройками OpenSSL по умолчанию handshake зависает.
        self.ssl_context = ssl.create_default_context()
        self.ssl_context.maximum_version = ssl.TLSVersion.TLSv1_2
        self.ssl_context.set_ciphers("DEFAULT@SECLEVEL=1")

    def call(self, controller: str, query: dict[str, Any] | None = None) -> dict[str, Any]:
        payload: dict[str, Any] = {"server_token": self.token, "controller": controller}
        if self.app_version:
            payload["app_version"] = self.app_version
        if query is not None:
            payload["query"] = query
        body = json.dumps(payload, ensure_ascii=False).encode("utf-8")
        request = urllib.request.Request(
            self.base_url,
            data=body,
            headers={"Content-Type": "application/json; charset=utf-8"},
            method="POST",
        )
        try:
            with urllib.request.urlopen(request, timeout=self.timeout, context=self.ssl_context) as response:
                raw = response.read()
        except urllib.error.HTTPError as exc:
            detail = exc.read().decode("utf-8", errors="replace")[:500]
            raise CrocotimeError(f"HTTP {exc.code} от Crocotime ({controller}): {detail}") from exc
        except urllib.error.URLError as exc:
            raise CrocotimeError(f"Не удалось подключиться к Crocotime ({controller}): {exc.reason}") from exc
        try:
            data = json.loads(raw.decode("utf-8"))
        except (UnicodeDecodeError, json.JSONDecodeError) as exc:
            raise CrocotimeError(f"Crocotime вернул некорректный JSON для {controller}") from exc
        if not isinstance(data, dict):
            raise CrocotimeError(f"Неожиданный ответ Crocotime для {controller}")
        if data.get("error"):
            raise CrocotimeError(f"Ошибка Crocotime ({controller}): {data['error']}")
        result = data.get("result")
        if not isinstance(result, dict):
            # Некоторые версии сервера возвращают items/activities
            # на верхнем уровне, без обёртки result.
            result = data
        return result


@dataclass
class Employee:
    employee_id: int
    name: str
    department_id: int | None
    department: str


@dataclass
class DailyRow:
    employee_id: int
    employee: str
    department: str
    day: date
    norm: int
    worked: int
    productive: int
    distractions: int
    without_computer: int
    begin: int
    end: int
    status: str
    occupancy: list[float]


def resolve_hermes_setting(name: str) -> str | None:
    """Resolve a value from the process environment or active Hermes profile."""
    value = os.environ.get(name)
    if value and value.strip():
        return value.strip()
    try:
        result = subprocess.run(
            ["hermes", "config", "get", name],
            capture_output=True,
            text=True,
            encoding="utf-8",
            errors="replace",
            timeout=15,
            check=False,
        )
    except (FileNotFoundError, subprocess.TimeoutExpired):
        return None
    if result.returncode != 0:
        return None
    value = result.stdout.strip()
    return value or None


def parse_month(value: str) -> tuple[int, int]:
    try:
        parsed = datetime.strptime(value, "%Y-%m")
    except ValueError as exc:
        raise argparse.ArgumentTypeError("месяц должен быть в формате YYYY-MM") from exc
    return parsed.year, parsed.month


def month_bounds(year: int, month: int, tz: ZoneInfo) -> tuple[datetime, datetime]:
    start = datetime(year, month, 1, tzinfo=tz)
    if month == 12:
        end = datetime(year + 1, 1, 1, tzinfo=tz)
    else:
        end = datetime(year, month + 1, 1, tzinfo=tz)
    return start, end


def iter_days(start: datetime, end: datetime):
    cursor = start
    while cursor < end:
        nxt = cursor + timedelta(days=1)
        yield cursor, nxt
        cursor = nxt


def items(result: dict[str, Any], key: str = "items") -> list[dict[str, Any]]:
    value = result.get(key, [])
    return [x for x in value if isinstance(x, dict)] if isinstance(value, list) else []


def safe_int(value: Any, default: int = 0) -> int:
    try:
        return int(value)
    except (TypeError, ValueError):
        return default


def extract_departments(tree_items: list[dict[str, Any]]) -> tuple[dict[int, str], dict[int, set[int]]]:
    names: dict[int, str] = {}
    members: dict[int, set[int]] = defaultdict(set)

    def walk(node: dict[str, Any], ancestors: tuple[int, ...] = ()) -> set[int]:
        department_id = node.get("department_id")
        current_ancestors = ancestors
        if department_id is not None:
            dep_id = safe_int(department_id, -1)
            if dep_id >= 0:
                names[dep_id] = str(node.get("display_name") or f"Департамент {dep_id}")
                current_ancestors = ancestors + (dep_id,)
        found: set[int] = set()
        employee_id = node.get("employee_id")
        if employee_id is not None:
            emp_id = safe_int(employee_id, -1)
            if emp_id >= 0:
                found.add(emp_id)
                for dep_id in current_ancestors:
                    members[dep_id].add(emp_id)
        children = node.get("items", [])
        if isinstance(children, list):
            for child in children:
                if isinstance(child, dict):
                    found.update(walk(child, current_ancestors))
        if department_id is not None:
            dep_id = safe_int(department_id, -1)
            if dep_id >= 0:
                members[dep_id].update(found)
        return found

    for root in tree_items:
        walk(root)
    return names, members


def select_employees(
    employee_items: list[dict[str, Any]],
    department_names: dict[int, str],
    department_members: dict[int, set[int]],
    employee_filter: set[int],
    department_filter: set[int],
) -> list[Employee]:
    allowed_by_department: set[int] = set()
    for dep_id in department_filter:
        allowed_by_department.update(department_members.get(dep_id, set()))
    selected: list[Employee] = []
    for item in employee_items:
        emp_id = safe_int(item.get("employee_id"), -1)
        if emp_id < 0 or safe_int(item.get("is_deleted"), 0) == 1 or safe_int(item.get("is_enabled"), 1) == 0:
            continue
        if employee_filter and emp_id not in employee_filter:
            continue
        if department_filter and emp_id not in allowed_by_department:
            continue
        dep_id_raw = item.get("parent_group_id")
        dep_id = safe_int(dep_id_raw, -1) if dep_id_raw is not None else None
        if dep_id is not None and dep_id < 0:
            dep_id = None
        name = str(item.get("display_name") or " ".join(filter(None, [item.get("second_name"), item.get("first_name")])) or f"Сотрудник {emp_id}")
        selected.append(Employee(emp_id, name, dep_id, department_names.get(dep_id, "Без департамента")))
    selected.sort(key=lambda employee: (employee.department.casefold(), employee.name.casefold()))
    return selected


def status_from_schedule(item: dict[str, Any]) -> str:
    statuses = []
    mapping = (
        ("vacation_time", "Отпуск"),
        ("sick_time", "Больничный"),
        ("day_off_time", "Отгул"),
        ("holiday_time", "Праздник"),
    )
    for key, label in mapping:
        seconds = safe_int(item.get(key), 0)
        if seconds > 0:
            statuses.append(f"{label} {format_seconds_text(seconds)}")
    return ", ".join(statuses)


def format_seconds_text(seconds: int) -> str:
    sign = "-" if seconds < 0 else ""
    seconds = abs(int(seconds))
    hours, remainder = divmod(seconds, 3600)
    minutes = remainder // 60
    if hours and minutes:
        return f"{sign}{hours} ч {minutes:02d} мин"
    if hours:
        return f"{sign}{hours} ч"
    return f"{sign}{minutes} мин"


def occupancy_bins(day_start: datetime, intervals: list[tuple[int, int]], bin_count: int = 48) -> list[float]:
    day_epoch = int(day_start.timestamp())
    bin_size = 86400 // bin_count
    values: list[float] = []
    for index in range(bin_count):
        left = day_epoch + index * bin_size
        right = left + bin_size
        covered = 0
        for start, end in intervals:
            covered += max(0, min(right, end) - max(left, start))
        values.append(min(1.0, covered / bin_size))
    return values


def fetch_report_data(
    client: CrocotimeClient,
    year: int,
    month: int,
    tz: ZoneInfo,
    employee_filter: set[int],
    department_filter: set[int],
) -> tuple[list[Employee], dict[int, dict[str, Any]], list[DailyRow]]:
    employee_result = client.call("api_employees")
    department_result = client.call("api_departments")
    dep_names, dep_members = extract_departments(items(department_result))
    employees = select_employees(items(employee_result), dep_names, dep_members, employee_filter, department_filter)
    if not employees:
        raise CrocotimeError("После применения фильтров не осталось сотрудников")
    employee_ids = [employee.employee_id for employee in employees]
    employee_by_id = {employee.employee_id: employee for employee in employees}
    start, end = month_bounds(year, month, tz)
    month_result = client.call(
        "api_employee_activity",
        {"interval": [int(start.timestamp()), int(end.timestamp())], "employees": employee_ids},
    )
    monthly = {safe_int(item.get("employee_id"), -1): item for item in items(month_result)}

    train_by_employee: dict[int, list[tuple[int, int]]] = defaultdict(list)
    for employee in employees:
        result = client.call(
            "api_window_switch_train",
            {"interval": [int(start.timestamp()), int(end.timestamp())], "employee_id": employee.employee_id},
        )
        for activity in items(result, "activities"):
            interval = activity.get("interval")
            if isinstance(interval, list) and len(interval) == 2:
                left, right = safe_int(interval[0], -1), safe_int(interval[1], -1)
                if left >= 0 and right > left:
                    train_by_employee[employee.employee_id].append((left, right))

    daily_rows: list[DailyRow] = []
    # У некоторых версий сервера activity.norm по интервалу [00:00, 24:00]
    # захватывает соседний день расписания; авторитетный источник дневной
    # нормы — api_employee_working_day_schedule.
    schedule_norm_totals: dict[int, int] = defaultdict(int)
    for day_start, day_end in iter_days(start, end):
        query_interval = {"interval": [int(day_start.timestamp()), int(day_end.timestamp())], "employees": employee_ids}
        day_epoch = int(day_start.timestamp())
        activity_result = client.call("api_employee_activity", query_interval)
        periods_result = client.call("api_employee_work_periods", {"day": day_epoch, "employees": employee_ids})
        schedule_result = client.call("api_employee_working_day_schedule", {"day": day_epoch, "employees": employee_ids})
        activity_map = {safe_int(item.get("employee_id"), -1): item for item in items(activity_result)}
        periods_map = {safe_int(item.get("employee_id"), -1): item for item in items(periods_result)}
        schedule_map = {safe_int(item.get("employee_id"), -1): item for item in items(schedule_result)}
        for employee_id in employee_ids:
            activity = activity_map.get(employee_id, {})
            period = periods_map.get(employee_id, {})
            schedule = schedule_map.get(employee_id, {})
            worked = safe_int(activity.get("summary_time"), 0)
            productive = safe_int(activity.get("permitted_time"), 0)
            distractions = safe_int(activity.get("forbidden_time"), 0)
            without = safe_int(activity.get("unknown_time"), max(0, worked - productive - distractions))
            status = status_from_schedule(schedule)
            schedule_norm_totals[employee_id] += safe_int(schedule.get("norm"), 0)
            if worked <= 0 and not status:
                continue
            norm = safe_int(schedule.get("norm"), safe_int(activity.get("norm"), 0))
            begin = safe_int(period.get("begin"), -1)
            end_seconds = safe_int(period.get("end"), -1)
            employee = employee_by_id[employee_id]
            daily_rows.append(
                DailyRow(
                    employee_id=employee_id,
                    employee=employee.name,
                    department=employee.department,
                    day=day_start.date(),
                    norm=norm,
                    worked=worked,
                    productive=productive,
                    distractions=distractions,
                    without_computer=without,
                    begin=begin,
                    end=end_seconds,
                    status=status,
                    occupancy=occupancy_bins(day_start, train_by_employee[employee_id]),
                )
            )
    # Месячная норма — сумма дневных норм расписания: месячный activity.norm
    # может быть завышен на день из-за включительной границы интервала.
    for employee_id, norm_total in schedule_norm_totals.items():
        if employee_id in monthly and isinstance(monthly[employee_id], dict):
            monthly[employee_id]["norm"] = norm_total
    return employees, monthly, daily_rows


def duration_value(seconds: int) -> float:
    return seconds / 86400


def safe_sheet_name(name: str, used: set[str]) -> str:
    base = INVALID_SHEET_CHARS.sub("_", name).strip(" '") or "Сотрудник"
    base = base[:31]
    candidate = base
    counter = 2
    while candidate.casefold() in used:
        suffix = f" ({counter})"
        candidate = base[:31 - len(suffix)] + suffix
        counter += 1
    used.add(candidate.casefold())
    return candidate


def apply_font_and_border(ws, min_row: int, max_row: int, min_col: int, max_col: int):
    for row in ws.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col):
        for cell in row:
            if cell.font.name != "Arial":
                cell.font = Font(name="Arial", size=10, bold=cell.font.bold, italic=cell.font.italic, color=cell.font.color)
            cell.border = ROW_BORDER


def add_summary_sheet(wb: Workbook, title: str, employees: list[Employee], monthly: dict[int, dict[str, Any]]):
    ws = wb.active
    ws.title = title
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A2"
    headers = ["Сотрудник", "Отработанное", "Выполнение нормы", "Переработка", "Отвлечения", "Продуктивное", "Без компьютера", "Employee ID", "Норма"]
    ws.append(headers)
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(vertical="center")
    ws.row_dimensions[1].height = 27
    ws.column_dimensions["A"].width = 43
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 22
    ws.column_dimensions["E"].width = 18
    ws.column_dimensions["F"].width = 20
    ws.column_dimensions["G"].width = 22
    ws.column_dimensions["H"].hidden = True
    ws.column_dimensions["I"].hidden = True

    grouped: dict[str, list[Employee]] = defaultdict(list)
    for employee in employees:
        grouped[employee.department].append(employee)
    current_row = 2
    for department in sorted(grouped, key=str.casefold):
        group_row = current_row
        ws.cell(group_row, 1, department)
        ws.cell(group_row, 1).font = Font(name="Arial", size=10, bold=True)
        for col in range(1, 8):
            ws.cell(group_row, col).fill = GROUP_FILL
        current_row += 1
        start_employee_row = current_row
        for employee in grouped[department]:
            row = current_row
            ws.cell(row, 1, employee.name)
            ws.cell(row, 8, employee.employee_id)
            month_item = monthly.get(employee.employee_id, {})
            worked = safe_int(month_item.get("summary_time"), 0)
            norm = safe_int(month_item.get("norm"), 0)
            productive = safe_int(month_item.get("permitted_time"), 0)
            distractions = safe_int(month_item.get("forbidden_time"), 0)
            without = safe_int(month_item.get("unknown_time"), max(0, worked - productive - distractions))
            ws.cell(row, 9, duration_value(norm))
            ws.cell(row, 2, duration_value(worked))
            ws.cell(row, 3, f"=IFERROR(B{row}/I{row},0)" if norm else (1 if worked else 0))
            ws.cell(row, 4, f"=B{row}-I{row}")
            ws.cell(row, 5, duration_value(distractions))
            ws.cell(row, 6, duration_value(productive))
            ws.cell(row, 7, duration_value(without))
            current_row += 1
        end_employee_row = current_row - 1
        for col in (2, 4, 5, 6, 7):
            letter = get_column_letter(col)
            ws.cell(group_row, col, f"=SUM({letter}{start_employee_row}:{letter}{end_employee_row})")
        ws.cell(group_row, 9, f"=SUM(I{start_employee_row}:I{end_employee_row})")
        ws.cell(group_row, 3, f"=IFERROR(B{group_row}/I{group_row},0)")

    last_row = current_row - 1
    for row in range(2, last_row + 1):
        for col in (2, 4, 5, 6, 7):
            ws.cell(row, col).number_format = DURATION_FORMAT if col != 4 else SIGNED_DURATION_FORMAT
        ws.cell(row, 3).number_format = PERCENT_FORMAT
        ws.cell(row, 1).alignment = Alignment(vertical="center")
        for col in range(2, 8):
            ws.cell(row, col).alignment = Alignment(horizontal="right", vertical="center")
        ws.row_dimensions[row].height = 27
    if last_row >= 2:
        ws.conditional_formatting.add(f"C2:C{last_row}", DataBarRule(start_type="num", start_value=0, end_type="num", end_value=1.5, color="2BA323", showValue=True))
    apply_font_and_border(ws, 1, max(1, last_row), 1, 8)
    ws.auto_filter.ref = f"A1:G{max(1, last_row)}"
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.print_area = f"A1:G{max(1, last_row)}"


def add_employee_sheet(wb: Workbook, employee: Employee, rows: list[DailyRow], period_title: str, used: set[str]):
    short_name = employee.name
    sheet_name = safe_sheet_name(short_name, used)
    ws = wb.create_sheet(sheet_name)
    ws.sheet_view.showGridLines = False
    ws.merge_cells("A1:BH1")
    ws["A1"] = employee.name
    ws["A1"].font = Font(name="Arial", size=14, bold=True, color="2E5E8C")
    ws["A1"].alignment = Alignment(vertical="center")
    ws.merge_cells("A2:BH2")
    ws["A2"] = f"{employee.department} · {period_title}"
    ws["A2"].font = Font(name="Arial", size=9, color="6F7F8F")

    headers = ["Дата", "День", "Приход", "Уход", "Статус", "Норма", "Отработано", "Выполнение", "Переработка", "Отвлечения", "Продуктивное", "Без компьютера"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(4, col, header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    timeline_start = 13
    for hour in range(24):
        first = timeline_start + hour * 2
        ws.merge_cells(start_row=4, start_column=first, end_row=4, end_column=first + 1)
        cell = ws.cell(4, first, f"{hour:02d}")
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
    widths = [12, 7, 9, 9, 18, 13, 15, 13, 16, 15, 17, 20]
    for col, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = width
    for col in range(timeline_start, timeline_start + 48):
        ws.column_dimensions[get_column_letter(col)].width = 2.1
    ws.freeze_panes = "M5"
    ws.row_dimensions[4].height = 31

    for index, daily in enumerate(sorted(rows, key=lambda row: row.day), 5):
        ws.cell(index, 1, daily.day)
        ws.cell(index, 2, WEEKDAYS_RU[daily.day.weekday()])
        ws.cell(index, 3, duration_value(daily.begin) if daily.begin >= 0 else None)
        ws.cell(index, 4, duration_value(daily.end) if daily.end >= 0 else None)
        ws.cell(index, 5, daily.status)
        ws.cell(index, 6, duration_value(daily.norm))
        ws.cell(index, 7, duration_value(daily.worked))
        ws.cell(index, 8, f"=IFERROR(G{index}/F{index},IF(G{index}>0,1,0))")
        ws.cell(index, 9, f"=G{index}-F{index}")
        ws.cell(index, 10, duration_value(daily.distractions))
        ws.cell(index, 11, duration_value(daily.productive))
        ws.cell(index, 12, duration_value(daily.without_computer))
        ws.cell(index, 1).number_format = "dd.mm.yyyy"
        for col in (3, 4):
            ws.cell(index, col).number_format = TIME_FORMAT
        for col in (6, 7, 10, 11, 12):
            ws.cell(index, col).number_format = DURATION_FORMAT
        ws.cell(index, 8).number_format = PERCENT_FORMAT
        ws.cell(index, 9).number_format = SIGNED_DURATION_FORMAT
        ws.cell(index, 5).alignment = Alignment(wrap_text=True, vertical="center")
        for bin_index, occupancy in enumerate(daily.occupancy):
            col = timeline_start + bin_index
            cell = ws.cell(index, col, occupancy if occupancy > 0 else None)
            cell.number_format = ";;;"
            bin_left = bin_index * 1800
            bin_right = bin_left + 1800
            in_span = daily.begin >= 0 and daily.end >= 0 and bin_right > daily.begin and bin_left < daily.end
            if occupancy >= 0.5:
                cell.fill = GREEN_FILL
            elif occupancy > 0:
                cell.fill = LIGHT_GREEN_FILL
            elif in_span:
                cell.fill = GRAY_FILL
            else:
                cell.fill = WHITE_FILL
        ws.row_dimensions[index].height = 25

    last_row = max(5, 4 + len(rows))
    apply_font_and_border(ws, 4, last_row, 1, 60)
    ws.auto_filter.ref = f"A4:L{last_row}"
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.print_area = f"A1:BH{last_row}"
    ws.print_title_rows = "1:4"


def add_data_sheet(wb: Workbook, rows: list[DailyRow], source_url: str, period_title: str):
    ws = wb.create_sheet("Данные")
    headers = ["Employee ID", "Департамент", "Сотрудник", "Дата", "Норма", "Отработано", "Продуктивное", "Отвлечения", "Без компьютера", "Приход", "Уход", "Статус"]
    ws.append(headers)
    for daily in sorted(rows, key=lambda row: (row.employee.casefold(), row.day)):
        ws.append([
            daily.employee_id,
            daily.department,
            daily.employee,
            daily.day,
            duration_value(daily.norm),
            duration_value(daily.worked),
            duration_value(daily.productive),
            duration_value(daily.distractions),
            duration_value(daily.without_computer),
            duration_value(daily.begin) if daily.begin >= 0 else None,
            duration_value(daily.end) if daily.end >= 0 else None,
            daily.status,
        ])
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
    for row in range(2, ws.max_row + 1):
        ws.cell(row, 4).number_format = "dd.mm.yyyy"
        for col in range(5, 10):
            ws.cell(row, col).number_format = DURATION_FORMAT
        for col in (10, 11):
            ws.cell(row, col).number_format = TIME_FORMAT
    ws.auto_filter.ref = ws.dimensions
    ws.freeze_panes = "A2"
    ws.sheet_state = "hidden"

    meta = wb.create_sheet("Параметры")
    meta.append(["Параметр", "Значение"])
    meta.append(["Период", period_title])
    meta.append(["Источник", source_url])
    meta.append(["Сформировано", datetime.now().astimezone().replace(tzinfo=None)])
    meta.append(["Примечание", "Токен Crocotime в файл не записывается"])
    meta["B4"].number_format = "dd.mm.yyyy hh:mm"
    meta.sheet_state = "hidden"


def build_workbook(
    output: Path,
    year: int,
    month: int,
    employees: list[Employee],
    monthly: dict[int, dict[str, Any]],
    daily_rows: list[DailyRow],
    source_url: str,
):
    wb = Workbook()
    month_title = MONTHS_RU[month]
    period_title = f"{MONTHS_RU[month].lower()} {year}"
    add_summary_sheet(wb, month_title, employees, monthly)
    used = {month_title.casefold()}
    rows_by_employee: dict[int, list[DailyRow]] = defaultdict(list)
    for row in daily_rows:
        rows_by_employee[row.employee_id].append(row)
    for employee in employees:
        add_employee_sheet(wb, employee, rows_by_employee[employee.employee_id], period_title, used)
    add_data_sheet(wb, daily_rows, source_url, period_title)
    wb.calculation.fullCalcOnLoad = True
    wb.calculation.forceFullCalc = True
    wb.calculation.calcMode = "auto"
    output.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output)


def verify_workbook(path: Path, expected_employees: int) -> dict[str, Any]:
    wb = load_workbook(path, data_only=False)
    formula_count = 0
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                if isinstance(cell.value, str) and cell.value.startswith("="):
                    formula_count += 1
    visible = [ws.title for ws in wb.worksheets if ws.sheet_state == "visible"]
    wb.close()
    if len(visible) != expected_employees + 1:
        raise CrocotimeError(f"Проверка XLSX не пройдена: ожидалось {expected_employees + 1} видимых листов, найдено {len(visible)}")
    return {"visible_sheets": visible, "formulas": formula_count, "size_bytes": path.stat().st_size}


def parse_args(argv: list[str]) -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Сформировать месячный XLSX-отчёт по данным Crocotime")
    parser.add_argument("--month", required=True, help="месяц YYYY-MM")
    parser.add_argument("--output", required=True, type=Path, help="путь к итоговому .xlsx")
    parser.add_argument("--base-url", help="URL Crocotime; по умолчанию CROCOTIME_URL из активного профиля Hermes")
    parser.add_argument("--token-env", default="CROCOTIME_TOKEN", help="имя настройки/переменной с токеном")
    parser.add_argument("--app-version", help="версия API/Crocotime; по умолчанию CROCOTIME_APP_VERSION")
    parser.add_argument("--timezone", default="Europe/Moscow", help="часовой пояс отчёта")
    parser.add_argument("--employee-id", type=int, action="append", default=[], help="ID сотрудника; можно повторять")
    parser.add_argument("--department-id", type=int, action="append", default=[], help="ID департамента; включает вложенных сотрудников")
    parser.add_argument("--timeout", type=int, default=60, help="таймаут одного HTTP-запроса")
    return parser.parse_args(argv)


def main(argv: list[str] | None = None) -> int:
    args = parse_args(argv or sys.argv[1:])
    year, month = parse_month(args.month)
    base_url = args.base_url or resolve_hermes_setting("CROCOTIME_URL")
    if not base_url:
        raise CrocotimeError("Не задан URL: укажите --base-url или настройте CROCOTIME_URL в Hermes")
    token = resolve_hermes_setting(args.token_env)
    if not token:
        raise CrocotimeError(f"Не задан токен: настройте {args.token_env} в активном профиле Hermes")
    app_version = args.app_version or resolve_hermes_setting("CROCOTIME_APP_VERSION")
    try:
        tz = ZoneInfo(args.timezone)
    except Exception as exc:
        raise CrocotimeError(f"Неизвестный часовой пояс: {args.timezone}") from exc
    client = CrocotimeClient(base_url, token, app_version, args.timeout)
    employees, monthly, daily_rows = fetch_report_data(
        client,
        year,
        month,
        tz,
        set(args.employee_id),
        set(args.department_id),
    )
    build_workbook(args.output, year, month, employees, monthly, daily_rows, base_url)
    verification = verify_workbook(args.output, len(employees))
    print(json.dumps({"status": "success", "output": str(args.output.resolve()), "employees": len(employees), "daily_rows": len(daily_rows), **verification}, ensure_ascii=False))
    return 0


if __name__ == "__main__":
    try:
        raise SystemExit(main())
    except CrocotimeError as exc:
        print(json.dumps({"status": "error", "error": str(exc)}, ensure_ascii=False), file=sys.stderr)
        raise SystemExit(2)
